# Filename: RateComparison_Windows.py
# Author: Zheng Guo
# Date: 10-07-2020
# Purpose: Comparing the current rate from ClaimSphere measure rate report with the previous year's rate and the submitted rate.
# Class list: - measure (name, rate, denominator, numerator)

# Functions:
# - each_population(loc,month)
# - write_diff_1(ws,month)
# - write_new(ws)
# - format_excel(loc, data)

# User Input:
# - First input: RC_FileName.xlsx
#                xlsx file that contains the follow information in this order:
#                total ppo now, total ppo pre, total ppo sub, total hmo now, total hmo pre, total hmo sub, fep now, fep pre, fep sub
#                B[2],B[3],B[4],B[5],B[6],B[7],B[8],B[9],B[10]
# - Second input: Current year's ClaimSphere measaure rate report
# - Third input: Previous year's ClaimSphere measaure rate report
# - Fourth input: Submitted ClaimSphere measaure rate report
# - Fifth input: The current month's abbreviation

################################################################################################################################################
# Imports
import decimal
import os
from datetime import date

import openpyxl
import pandas as pd
import xlrd
import xlsxwriter
import xlwt
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side, fills
from xlwt.Style import *


################################################################################################################################################
# Class
class Measure():
    def __init__(self, name, rate, denominator, numerator):
        self.name = name
        self.rate = rate
        self.denominator = denominator
        self.numerator = numerator

################################################################################################################################################
# Functions


def each_population(loc, month):
    loc_new = loc[0]
    loc_old = loc[1]
    loc_sub = loc[2]

    # Sub Functions
    def write_diff_1(ws, month):
        ws.write(0, 0, "SubMeasureID", style)
        ws.write(0, 1, month+"19_Rate", style)
        ws.write(0, 2, month+"20_Rate", style)
        ws.write(0, 3, "DifferenceI", style)
        ws.write(0, 4, "Sub20_Rate", style)
        ws.write(0, 5, "DifferenceII", style)

        ws.write(0, 6, "------------", style)
        ws.write(0, 7, month+"19_Denom", style)
        ws.write(0, 8, month+"20_Denom", style)
        ws.write(0, 9, "Diff_Denom", style)

        ws.write(0, 10, "-------------", style)
        ws.write(0, 11, month+"19_Num", style)
        ws.write(0, 12, month+"20_Num", style)
        ws.write(0, 13, "Diff_Num", style)
        the_measure.sort()

        i = 0
        while (i < len(the_measure)):
            the_measure_name = the_measure[i]

            old_rate = "a"
            old_denominator = "a"
            old_numerator = "a"
            new_rate = "a"
            new_denominator = "a"
            new_numerator = "a"
            sub_rate = "a"
            sub_denominator = "a"
            sub_numerator = "a"
            old = old_measure[the_measure_name]
            old_rate = old.rate
            old_denominator = old.denominator
            old_numerator = old.numerator
            new = new_measure[the_measure_name]
            new_rate = new.rate
            new_denominator = new.denominator
            new_numerator = new.numerator

            ws.write(i+1, 0, the_measure_name, style)
            ws.write(i+1, 1, old_rate, style)
            ws.write(i+1, 7, old_denominator, style)
            ws.write(i+1, 11, old_numerator, style)
            ws.write(i+1, 2, new_rate, style)
            ws.write(i+1, 8, new_denominator, style)
            ws.write(i+1, 12, new_numerator, style)

            new_denominator = new_denominator.replace(',', '')
            old_denominator = old_denominator.replace(',', '')
            new_numerator = new_numerator.replace(',', '')
            old_numerator = old_numerator.replace(',', '')

            ws.write(i+1, 3, decimal.Decimal(new_rate) -
                     decimal.Decimal(old_rate), style)
            ws.write(i+1, 9, str(int(new_denominator) -
                                 int(old_denominator)), style)
            ws.write(i+1, 13, str(int(new_numerator)-int(old_numerator)), style)

            i = i+1

    def write_new(ws):
        i = 0
        n = len(the_measure)
        while i < count:
            new_measure_name = new_list[i]
            ws.write(n+1, 0, new_measure_name, style)
            new_rate = "a"
            new_denominator = "a"
            new_numerator = "a"
            new = new_measure[new_measure_name]
            new_rate = new.rate
            new_denominator = new.denominator
            new_numerator = new.numerator
            ws.write(n+1, 2, new_rate, style)
            ws.write(n+1, 8, new_denominator, style)
            ws.write(n+1, 12, new_numerator, style)
            i = i+1
            n = n+1

        book.save(f'{Population_name}Difference.xlsx')
        workbook = xlrd.open_workbook(f'{Population_name}Difference.xlsx')
        worksheet = workbook.sheet_by_index(0)

        j = 1
        while j < worksheet.nrows:
            measure_id = worksheet.cell(j, 0).value
            data.append(worksheet.cell(j, 3).value)
            sub = sub_measure[measure_id]
            sub_rate = sub.rate
            new = new_measure[measure_id]
            new_rate = new.rate
            ws.write(j, 4, sub_rate, style)
            ws.write(j, 5, decimal.Decimal(new_rate) -
                     decimal.Decimal(sub_rate), style)
            ws.write(j, 6, '', style)
            ws.write(j, 10, '', style)
            j = j+1

        book.save(f'{Population_name}Difference.xlsx')

    def format_excel(loc, data):
        df = pd.read_excel(loc)
        df['DifferenceI'] = df['DifferenceI'].astype(float)
        df['DifferenceII'] = df['DifferenceII'].astype(float)
        df['Diff_Denom'] = df['Diff_Denom'].astype(float)
        df['Diff_Num'] = df['Diff_Num'].astype(float)
        df.style.bar(subset=['B', 'C'], color='#d65f5f')
        threahold = 0.0
        df_styled = df.style\
            .applymap(lambda x: 'background-color: %s' % 'pink'if x < threahold else 'background-color: %s' % 'white', subset=['DifferenceI'])\
            .applymap(lambda x: 'background-color: %s' % 'pink' if x < threahold else 'background-color: %s' % 'white', subset=['DifferenceII'])\
            .applymap(lambda x: 'background-color: %s' % 'pink' if x < threahold else 'background-color: %s' % 'white', subset=['Diff_Denom'])\
            .applymap(lambda x: 'background-color: %s' % 'pink' if x < threahold else 'background-color: %s' % 'white', subset=['Diff_Num'])\
            .background_gradient(cmap='Blues', subset=['Sep20_Rate'])\
            .to_excel(f'{Population_name}_Difference_{now}.xlsx', engine='openpyxl', index=False)

    data = []
    old_measure = {}
    book_old = xlrd.open_workbook(loc_old)
    sheet_old = book_old.sheet_by_name("Sheet1")
    sheet_old.cell_value(0, 0)
    r_old = sheet_old.nrows
    name_row_old_start = 1
    name_row_old_end = r_old-1
    name_column_old = 3
    rate_column_old = 11
    denominator_column_old = 6
    numerator_column_old = 7

    new_measure = {}
    book_new = xlrd.open_workbook(loc_new)
    sheet_new = book_new.sheet_by_name("Sheet1")
    sheet_new.cell_value(0, 0)
    r_new = sheet_new.nrows
    name_row_new_start = 1
    name_row_new_end = r_new-1
    name_column_new = 3
    rate_column_new = 11
    denominator_column_new = 6
    numerator_column_new = 7

    sub_measure = {}
    book_sub = xlrd.open_workbook(loc_sub)
    sheet_sub = book_sub.sheet_by_name("Sheet1")
    sheet_sub.cell_value(0, 0)
    r_sub = sheet_sub.nrows
    name_row_sub_start = 0
    name_row_sub_end = r_sub-1
    name_column_sub = 5
    rate_column_sub = 13
    denominator_column_sub = 8
    numerator_column_sub = 9

    Population_name = sheet_sub.cell_value(1, 1)
    # assign measure(class) and name into a dictionary called old_measure
    i = name_row_old_start
    while (i < name_row_old_end):
        if str(sheet_old.cell_value(i, name_column_old)) == 'UOD':
            name = 'HDO'
        else:
            name = str(sheet_old.cell_value(i, name_column_old))
        rate = str(sheet_old.cell_value(i, rate_column_old))
        denominator = str(sheet_old.cell_value(i, denominator_column_old))
        numerator = str(sheet_old.cell_value(i, numerator_column_old))
        measure = Measure(name, rate, denominator, numerator)
        old_measure[name] = measure
        i = i+1

    # assign measure(class) and name into a dictionary called new_measure
    i = name_row_new_start
    while (i < name_row_new_end):
        name = str(sheet_new.cell_value(i, name_column_new))
        rate = str(sheet_new.cell_value(i, rate_column_new))
        denominator = str(sheet_new.cell_value(i, denominator_column_new))
        numerator = str(sheet_new.cell_value(i, numerator_column_new))
        measure = Measure(name, rate, denominator, numerator)
        new_measure[name] = measure
        i = i+1

    # assign measure(class) and name into a dictionary called sub_measure
    i = name_row_sub_start
    while (i < name_row_sub_end):
        name = str(sheet_sub.cell_value(i, name_column_sub))
        rate = str(sheet_sub.cell_value(i, rate_column_sub))
        denominator = str(sheet_sub.cell_value(i, denominator_column_sub))
        numerator = str(sheet_sub.cell_value(i, numerator_column_sub))
        measure = Measure(name, rate, denominator, numerator)
        sub_measure[name] = measure
        i = i+1

    old_measure_name = list(old_measure.keys())
    new_measure_name = list(new_measure.keys())
    sub_measure_name = list(sub_measure.keys())

    # assign the key of this dictionary into a list, a list of all new measure and a list of all old measure
    length_of_old_measure = len(old_measure_name)
    length_of_sub_measure = len(sub_measure_name)
    the_measure = []
    the_sub_measure = []

    i = 0
    while (i < length_of_old_measure):
        the_measure_name = old_measure_name[i]
        if (the_measure_name not in the_measure):
            if (the_measure_name in new_measure_name):
                the_measure.append(the_measure_name)
                # find out the shared measure between two list and assign those into a new list: the measure
        i = i+1

    book = xlwt.Workbook()
    ws = book.add_sheet("Differences")
    new_list = list(set(new_measure_name)-set(old_measure_name))
    new_list.sort()
    count = len(new_list)
    write_diff_1(ws, month)
    write_new(ws)
    cwd = os.getcwd()
    loc = f'{cwd}\{Population_name}Difference.xlsx'
    format_excel(loc, data)
    os.remove(f"{Population_name}Difference.xlsx")
    wb = openpyxl.load_workbook(f'{Population_name}_Difference_{now}.xlsx')
    sheet = wb['Sheet1']
    sheet.row_dimensions[1].height = 25
    sheet.column_dimensions['A'].width = 12
    sheet.column_dimensions['B'].width = 12
    sheet.column_dimensions['C'].width = 12
    sheet.column_dimensions['D'].width = 12
    sheet.column_dimensions['E'].width = 12
    sheet.column_dimensions['F'].width = 12
    sheet.column_dimensions['G'].width = 12
    sheet.column_dimensions['H'].width = 12
    sheet.column_dimensions['I'].width = 12
    sheet.column_dimensions['J'].width = 12
    sheet.column_dimensions['K'].width = 12
    sheet.column_dimensions['L'].width = 12
    sheet.column_dimensions['M'].width = 12
    sheet.column_dimensions['N'].width = 12
    sheet.freeze_panes = 'A2'

    row_num = 220
    col_num = 14
    row_loc = 0
    col_loc = 0

    for i in range(row_loc, row_num):
        for j in range(col_loc, col_num):
            sheet.cell(row=i+1, column=j+1).border = thin_border
            if i == row_loc:
                sheet.cell(row=i+1, column=j+1).border = thin_border
            if i == row_loc+row_num-1:
                sheet.cell(row=i+1, column=j+1).border = thin_border
    row_loc = row_loc+row_num

    wb.save(f'{Population_name}_Difference_{now}.xlsx')


################################################################################################################################################
print('\n Please make sure all measure report''s name are listed in the file RC_FileName.xlsx')
print('\n You will be prompted to type in the name of the month of this prospective run')
month = input('\n Ths abbreviation of the month is:')

my_cwd = os.getcwd()
style = xlwt.easyxf("align:vert centre,horiz right")
now = str(date.today()).replace('-', '_')

wb = load_workbook('RC_Filename.xlsx')
ws = wb.active
a = []

for row in ws.iter_rows(min_row=2, max_col=2, max_row=10):
    for cell in row:
        a.append(cell.value)

# windows system using backwards slash
ppo_now = my_cwd+'\\'+a[1]+'.xlsx'
ppo_pre = my_cwd+'\\'+a[3]+'.xlsx'
ppo_sub = my_cwd+'\\'+a[5]+'.xlsx'
hmo_now = my_cwd+'\\'+a[7]+'.xlsx'
hmo_pre = my_cwd+'\\'+a[9]+'.xlsx'
hmo_sub = my_cwd+'\\'+a[11]+'.xlsx'
fep_now = my_cwd+'\\'+a[13]+'.xlsx'
fep_pre = my_cwd+'\\'+a[15]+'.xlsx'
fep_sub = my_cwd+'\\'+a[17]+'.xlsx'

loc_list = [[ppo_now, ppo_pre, ppo_sub],
            [hmo_now, hmo_pre, hmo_sub],
            [fep_now, fep_pre, fep_sub]
            ]
thin_border = Border(left=Side(border_style='dashed', color='FF000000'),
                     right=Side(border_style='dashed', color='FF000000'),
                     top=Side(border_style='thin', color='FF000000'),
                     bottom=Side(border_style='thin', color='FF000000')
                     )

for loc in loc_list:
    each_population(loc, month)

print(' \n The script is completed')
